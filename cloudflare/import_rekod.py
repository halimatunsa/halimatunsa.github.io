# ============================================================
# IMPORT REKOD MURID DARIPADA FAIL EXCEL CIKGU
# Baca:
#   - _ANALISIS MARKAH PPT SAINS T5 2026.xlsx (sheet ANALISIS)
#       -> murid T5 2026 + 5 penilaian (UP/PPSA/PASA 2025, UP1/PPT 2026)
#   - Borang Rekod Transit (senarai nama) -> roster T3 & T4 + semakan T5
# Hasil:
#   - rekod_murid.json  (format sama dengan "rekod" dalam data.js)
# Guna terbit_rekod.py untuk hantar ke pangkalan data (D1/PHP).
# ============================================================
import openpyxl, os, json, re, sys

FOLDER = r"C:\Users\Fiz\Documents\cg matun"
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "rekod_murid.json")

KELAS_MAP = {"5M": "5 Ma'wa", "5A": "5 Adni", "5R": "5 Rayyan"}

def bersih(nama):
    return re.sub(r"\s+", " ", str(nama or "").strip())

def kelas_cantik(k):
    """'3 FIRDAUS' -> '3 Firdaus' (nombor kekal, perkataan jadi title case)"""
    return " ".join(w if w.isdigit() else w.capitalize() for w in bersih(k).split())

def norm(nama):
    return bersih(nama).lower()

def markah1dp(v):
    try:
        return round(float(v), 1)
    except (TypeError, ValueError):
        return None

rekod = {}          # norm(nama) -> rekod
pertindihan = []    # nama sama muncul lebih sekali

def tambah(nama, tingkatan, kelas, exam=None):
    k = norm(nama)
    if not k:
        return
    if k in rekod:
        pertindihan.append(f"{bersih(nama)} ({rekod[k]['kelas']} vs {kelas})")
        return
    rekod[k] = {
        "nama": bersih(nama),
        "ic": "",
        "tingkatan": tingkatan,
        "kelas": kelas,
        "contoh": False,
        "exam": exam or [],
        "pbd": [],
    }

# ---------- 1) Murid T5 2026 daripada sheet ANALISIS ----------
ppt_path = os.path.join(FOLDER, "_ANALISIS MARKAH PPT SAINS T5 2026.xlsx")
wb = openpyxl.load_workbook(ppt_path, data_only=True)
ws = wb["ANALISIS"]
# lajur: A BIL, B NAMA, C KELAS, D UP, E G, F PPSA, G G, H PASA, I G, J UP1, K G, L PPT, M G
n_t5 = 0
for r in range(3, ws.max_row + 1):
    nama = ws.cell(r, 2).value
    if not bersih(nama):
        break
    kelas_kod = bersih(ws.cell(r, 3).value)
    kelas = KELAS_MAP.get(kelas_kod, kelas_kod)
    exam = []
    for label, mcol in [("UP 2025", 4), ("PPSA 2025", 6), ("PASA 2025", 8), ("UP1 2026", 10), ("PPT 2026", 12)]:
        m = markah1dp(ws.cell(r, mcol).value)
        g = bersih(ws.cell(r, mcol + 1).value)
        if m is not None:
            exam.append({"nama": label, "markah": m, "gred": g or "-"})
    tambah(nama, "Tingkatan 5", kelas, exam)
    n_t5 += 1
wb.close()
print(f"T5 (ANALISIS): {n_t5} murid dibaca")

# ---------- 2) Roster daripada borang transit ----------
def baca_senarai(fail):
    """Pulangkan (kelas, [nama...]) daripada sheet 'Senarai nama'."""
    wb = openpyxl.load_workbook(os.path.join(FOLDER, fail), data_only=True)
    ws = wb["Senarai nama"]
    kelas = ""
    header_row = None
    for r in range(1, 15):
        a = bersih(ws.cell(r, 1).value).upper()
        if a.startswith("KELAS"):
            kelas = bersih(ws.cell(r, 3).value)
        if a == "BIL":
            header_row = r
            break
    nama_list = []
    if header_row:
        for r in range(header_row + 1, ws.max_row + 1):
            nm = bersih(ws.cell(r, 2).value)
            if not nm:
                break
            nama_list.append(nm)
    wb.close()
    return kelas, nama_list

TRANSIT = [
    ("Borang Rekod Transit Sains  3 FIRDAUS.xlsx", "Tingkatan 3"),
    ("Borang Rekod Transit Sains  3 RAYYAN.xlsx", "Tingkatan 3"),
    ("Borang Rekod Transit Sains T3.xlsx", "Tingkatan 3"),
    ("Borang Rekod Transit Sains 4 ADNI.xlsx", "Tingkatan 4"),
    ("Borang Rekod Transit Sains 4 RAYYAN.xlsx", "Tingkatan 4"),
    ("Borang Rekod Transit Sains T4.xlsx", "Tingkatan 4"),
    ("Borang Rekod Transit Sains 5 ADNI.xlsx", "Tingkatan 5"),
    ("Borang Rekod Transit Sains 5 RAYYAN.xlsx", "Tingkatan 5"),
    ("Borang Rekod Transit Sains T5.xlsx", "Tingkatan 5"),
    ("Borang Rekod Transit Sains T5 (1).xlsx", "Tingkatan 5"),
]

kelas_dilihat = set()
for fail, ting in TRANSIT:
    kelas, nama_list = baca_senarai(fail)
    kunci_kelas = (ting, kelas.upper())
    if kunci_kelas in kelas_dilihat:
        print(f"  langkau {fail} (kelas {kelas!r} sudah diproses)")
        continue
    kelas_dilihat.add(kunci_kelas)
    baru = 0
    for nm in nama_list:
        if norm(nm) not in rekod:
            tambah(nm, ting, kelas_cantik(kelas))
            baru += 1
    print(f"  {fail}: kelas={kelas!r}, {len(nama_list)} nama, {baru} baharu")

# ---------- 3) Simpan ----------
senarai = sorted(rekod.values(), key=lambda x: (x["tingkatan"], x["kelas"], x["nama"]))
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(senarai, f, ensure_ascii=False, indent=2)

print(f"\nJUMLAH: {len(senarai)} murid -> {OUT}")
from collections import Counter
for (t, k), n in sorted(Counter((x["tingkatan"], x["kelas"]) for x in senarai).items()):
    print(f"  {t:12s} {k:12s} {n} murid")
ada_exam = sum(1 for x in senarai if x["exam"])
print(f"  dengan markah peperiksaan: {ada_exam}")
if pertindihan:
    print("\nAMARAN nama berulang (dilangkau):")
    for p in pertindihan:
        print("  -", p)
